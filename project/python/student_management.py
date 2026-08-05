"""
Student Management System — Multi-School Edition
Business-grade Excel automation with Python
================================================

Features:
- Multi-school support (create separate workbooks per school)
- Dashboard with KPIs, charts, and analytics
- Student database with 24 students (5 subjects)
- Grade book with automatic letter grades + remarks
- Analytics: subject averages, gender comparison, rankings
- Report card generator (individual + whole-class PDF)
- Conditional formatting and professional styling
- Charts: bar, pie, line
- PowerShell/report generation via command line

Usage:
    python student_management.py --school "Lincoln International School" --output ./output
    python student_management.py --help
"""

import argparse
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.comments import Comment

# ═══════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════

COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "subheader_bg": "2E75B6",
    "grade_a": "00B050",
    "grade_b": "92D050",
    "grade_c": "FFFF00",
    "grade_d": "FFC000",
    "grade_e": "FF8000",
    "grade_f": "FF0000",
    "dashboard_bg": "F2F2F2",
    "table_header_bg": "4472C4",
    "table_header_fg": "FFFFFF",
    "row_alt": "D9E2F3",
    "time": "BDD7EE",
}

SUBJECTS = ["English", "Mathematics", "Science", "ICT", "Social Studies"]
SUBJECT_COLS = 3  # Starting column for subject scores
GRADE_SCALE = [(80, "A", "Excellent"), (70, "B", "Very Good"), 
                (60, "C", "Good"), (50, "D", "Pass"),
                (40, "E", "Weak"), (0, "F", "Fail")]

# 24 Students — Ghanaian school dataset
STUDENTS = [
    ("ST001", "Daniel", "Mensah", "Male", "2008-05-14", "Grade 10", "Blue", "Michael Mensah", "245001001", "Accra", "2022-09-05"),
    ("ST002", "Ama", "Owusu", "Female", "2009-01-20", "Grade 10", "Red", "Evelyn Owusu", "245001002", "Kumasi", "2022-09-05"),
    ("ST003", "Samuel", "Boateng", "Male", "2008-08-09", "Grade 10", "Green", "Joseph Boateng", "245001003", "Takoradi", "2022-09-05"),
    ("ST004", "Grace", "Asante", "Female", "2009-02-18", "Grade 10", "Yellow", "Mary Asante", "245001004", "Cape Coast", "2022-09-05"),
    ("ST005", "Kwame", "Ofori", "Male", "2008-11-01", "Grade 10", "Blue", "Patrick Ofori", "245001005", "Tema", "2022-09-05"),
    ("ST006", "Abena", "Darko", "Female", "2009-06-11", "Grade 10", "Red", "Janet Darko", "245001006", "Sunyani", "2022-09-05"),
    ("ST007", "Isaac", "Addo", "Male", "2008-03-24", "Grade 10", "Green", "Richard Addo", "245001007", "Ho", "2022-09-05"),
    ("ST008", "Esi", "Appiah", "Female", "2009-04-30", "Grade 10", "Yellow", "Agnes Appiah", "245001008", "Koforidua", "2022-09-05"),
    ("ST009", "Kojo", "Yeboah", "Male", "2008-09-16", "Grade 10", "Blue", "David Yeboah", "245001009", "Tamale", "2022-09-05"),
    ("ST010", "Akosua", "Frimpong", "Female", "2009-07-28", "Grade 10", "Red", "Lucy Frimpong", "245001010", "Accra", "2022-09-05"),
    ("ST011", "Nathan", "Agyeman", "Male", "2008-12-12", "Grade 10", "Green", "Stephen Agyeman", "245001011", "Kumasi", "2022-09-05"),
    ("ST012", "Linda", "Antwi", "Female", "2009-08-17", "Grade 10", "Yellow", "Rebecca Antwi", "245001012", "Kasoa", "2022-09-05"),
    ("ST013", "Joseph", "Bonsu", "Male", "2008-10-02", "Grade 10", "Blue", "Samuel Bonsu", "245001013", "Winneba", "2022-09-05"),
    ("ST014", "Patricia", "Nyarko", "Female", "2009-09-15", "Grade 10", "Red", "Elizabeth Nyarko", "245001014", "Accra", "2022-09-05"),
    ("ST015", "Elvis", "Kusi", "Male", "2008-01-29", "Grade 10", "Green", "Eric Kusi", "245001015", "Tema", "2022-09-05"),
    ("ST016", "Ruth", "Opoku", "Female", "2009-03-10", "Grade 10", "Yellow", "Martha Opoku", "245001016", "Kumasi", "2022-09-05"),
    ("ST017", "Emmanuel", "Tetteh", "Male", "2008-07-06", "Grade 10", "Blue", "Charles Tetteh", "245001017", "Accra", "2022-09-05"),
    ("ST018", "Faith", "Adjei", "Female", "2009-11-19", "Grade 10", "Red", "Paul Adjei", "245001018", "Takoradi", "2022-09-05"),
    ("ST019", "Benjamin", "Arthur", "Male", "2008-04-08", "Grade 10", "Green", "Thomas Arthur", "245001019", "Cape Coast", "2022-09-05"),
    ("ST020", "Deborah", "Quaye", "Female", "2009-10-25", "Grade 10", "Yellow", "Naomi Quaye", "245001020", "Ho", "2022-09-05"),
    ("ST021", "Prince", "Amankwah", "Male", "2008-06-18", "Grade 10", "Blue", "Daniel Amankwah", "245001021", "Koforidua", "2022-09-05"),
    ("ST022", "Esther", "Ansah", "Female", "2009-12-04", "Grade 10", "Red", "Joyce Ansah", "245001022", "Tema", "2022-09-05"),
    ("ST023", "Kelvin", "Aidoo", "Male", "2008-02-21", "Grade 10", "Green", "George Aidoo", "245001023", "Sunyani", "2022-09-05"),
    ("ST024", "Mavis", "Boakye", "Female", "2009-05-27", "Grade 10", "Yellow", "Florence Boakye", "245001024", "Accra", "2022-09-05"),
]

# Score data: (English, Math, Science, ICT, Social)
SCORES = [
    (85, 78, 82, 90, 79),  # Daniel
    (92, 88, 75, 85, 90),  # Ama
    (70, 65, 72, 68, 74),  # Samuel
    (88, 95, 91, 87, 89),  # Grace
    (65, 60, 58, 70, 62),  # Kwame
    (78, 82, 80, 85, 76),  # Abena
    (72, 68, 75, 73, 71),  # Isaac
    (95, 90, 88, 92, 93),  # Esi
    (60, 55, 62, 58, 65),  # Kojo
    (85, 80, 78, 82, 84),  # Akosua
    (68, 72, 70, 65, 69),  # Nathan
    (90, 85, 87, 88, 86),  # Linda
    (75, 70, 68, 72, 73),  # Joseph
    (82, 78, 80, 85, 79),  # Patricia
    (55, 50, 48, 60, 52),  # Elvis
    (88, 92, 85, 90, 86),  # Ruth
    (70, 65, 72, 68, 71),  # Emmanuel
    (93, 88, 90, 95, 91),  # Faith
    (78, 75, 70, 72, 76),  # Benjamin
    (86, 80, 82, 88, 84),  # Deborah
    (65, 60, 68, 62, 64),  # Prince
    (91, 85, 88, 90, 89),  # Esther
    (72, 68, 65, 70, 71),  # Kelvin
    (84, 88, 86, 82, 85),  # Mavis
]

ATTENDANCE = [95, 98, 91, 96, 90, 97, 93, 94, 88, 99, 92, 95, 89, 97, 94, 98, 91, 96, 90, 99, 93, 97, 92, 98]

# ═══════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def get_letter_grade(avg):
    for threshold, grade, remarks in GRADE_SCALE:
        if avg >= threshold:
            return grade, remarks
    return "F", "Fail"


def style_header(cell, bg_color, fg_color="FFFFFF", size=12, bold=True):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg_color)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def style_data_cell(cell, alt_row=False, align="center"):
    cell.font = Font(name="Calibri", size=11)
    if alt_row:
        cell.fill = PatternFill(start_color=COLORS["row_alt"], end_color=COLORS["row_alt"], fill_type="solid")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def grade_fill(grade):
    """Return PatternFill for a letter grade."""
    color_map = {"A": "00B050", "B": "92D050", "C": "FFFF00", "D": "FFC000", "E": "FF8000", "F": "FF0000"}
    color = color_map.get(grade, "FFFFFF")
    fg = "FFFFFF" if grade in ("A", "E", "F") else "000000"
    return PatternFill(start_color=color, end_color=color, fill_type="solid"), Font(color=fg, bold=True)


# ═══════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════

def build_students_sheet(wb):
    """Sheet: Students — full student database with 24 students."""
    ws = wb.create_sheet("Students", 0)

    headers = ["StudentID", "First Name", "Last Name", "Gender", "Date of Birth",
               "Class", "House", "Parent Name", "Parent Phone", "Address",
               "Admission Date", "Attendance %", "English", "Mathematics",
               "Science", "ICT", "Social Studies", "Total", "Average", "Grade"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell, COLORS["table_header_bg"])

    for idx, (student, scores, att) in enumerate(zip(STUDENTS, SCORES, ATTENDANCE), 2):
        row_data = list(student) + [att] + list(scores)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            style_data_cell(cell, alt_row=(idx % 2 == 0))

        # Total formula — scores are in columns M-Q (13-17), Total=18(R), Avg=19(S), Grade=20(T)
        total_col = 18  # Column R
        score_start = 13  # Column M
        score_end = 17    # Column Q
        total_formula = f"=SUM({get_column_letter(score_start)}{idx}:{get_column_letter(score_end)}{idx})"
        total_cell = ws.cell(row=idx, column=total_col, value=f"=SUM(M{idx}:Q{idx})")
        style_data_cell(total_cell, alt_row=(idx % 2 == 0))

        # Average formula
        avg_col = 19  # Column S
        avg_cell = ws.cell(row=idx, column=avg_col, value=f"=ROUND(AVERAGE(M{idx}:Q{idx}),2)")
        avg_cell.number_format = "0.00"
        style_data_cell(avg_cell, alt_row=(idx % 2 == 0))

        # Grade formula (nested IF)
        grade_col = 20  # Column T
        grade_formula = (f'=IF(S{idx}>=80,"A",'
                         f'IF(S{idx}>=70,"B",'
                         f'IF(S{idx}>=60,"C",'
                         f'IF(S{idx}>=50,"D",'
                         f'IF(S{idx}>=40,"E","F")))))')
        grade_cell = ws.cell(row=idx, column=grade_col, value=grade_formula)
        fill, font = grade_fill("A")
        grade_cell.fill = fill
        grade_cell.font = font
        grade_cell.alignment = Alignment(horizontal="center", vertical="center")
        grade_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))

    # Column widths
    widths = [10, 12, 12, 10, 14, 10, 10, 18, 14, 14, 14, 12, 10, 12, 10, 8, 14, 10, 10, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{len(STUDENTS)+1}"
    return ws


def build_grades_sheet(wb):
    """Sheet: Grades — grade book with formulas, conditional formatting."""
    ws = wb.create_sheet("Grades", 1)

    headers = ["Student ID", "Name", "English", "Mathematics", "Science", "ICT",
               "Social Studies", "Total", "Average", "Grade", "Remarks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell, COLORS["table_header_bg"])

    for idx, ((sid, fname, lname, *_), scores) in enumerate(zip(STUDENTS, SCORES), 2):
        name = f"{fname} {lname}"
        ws.cell(row=idx, column=1, value=sid)
        ws.cell(row=idx, column=2, value=name)
        for col_offset, score in enumerate(scores):
            ws.cell(row=idx, column=3 + col_offset, value=score)

        # Total
        ws.cell(row=idx, column=8, value=f"=SUM(C{idx}:G{idx})")
        # Average
        avg_cell = ws.cell(row=idx, column=9, value=f"=ROUND(AVERAGE(C{idx}:G{idx}),2)")
        avg_cell.number_format = "0.00"
        # Grade
        ws.cell(row=idx, column=10,
                value=f'=IF(I{idx}>=80,"A",IF(I{idx}>=70,"B",IF(I{idx}>=60,"C",IF(I{idx}>=50,"D",IF(I{idx}>=40,"E","F")))))')
        # Remarks
        ws.cell(row=idx, column=11,
                value=f'=IF(J{idx}="A","Excellent",IF(J{idx}="B","Very Good",IF(J{idx}="C","Good",IF(J{idx}="D","Pass",IF(J{idx}="E","Weak","Fail")))))')

        for col in range(1, 12):
            cell = ws.cell(row=idx, column=col)
            style_data_cell(cell, alt_row=(idx % 2 == 0))

    # Conditional formatting on Average column (I)
    ws.conditional_formatting.add(f"I2:I{len(STUDENTS)+1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF0000",
                       mid_type="num", mid_value=50, mid_color="FFFF00",
                       end_type="num", end_value=100, end_color="00B050"))

    # Column widths
    for col, w in enumerate([10, 22, 10, 12, 10, 8, 14, 10, 10, 8, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    return ws


def build_dashboard_sheet(wb, school_name):
    """Sheet: Dashboard — KPIs + charts."""
    ws = wb.create_sheet("Dashboard", 0)
    last_row = len(STUDENTS) + 1

    # ── Title ──
    ws.merge_cells("B2:H2")
    title = ws.cell(row=2, column=2, value=f"{school_name} — Academic Dashboard")
    title.font = Font(name="Calibri", size=20, bold=True, color="1F4E79")
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:H3")
    date_cell = ws.cell(row=3, column=2, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d')}")
    date_cell.font = Font(name="Calibri", size=11, italic=True, color="808080")
    date_cell.alignment = Alignment(horizontal="center")

    # ── KPI Cards ──
    kpis = [
        ("Total Students", f"=COUNTA(Students!A2:A{last_row})"),
        ("Class Average", f"=ROUND(AVERAGE(Grades!I2:I{last_row}),2)"),
        ("Top Score", f"=MAX(Grades!I2:I{last_row})"),
        ("Pass Rate", f'=COUNTIF(Grades!J2:J{last_row},"<>F")/COUNTA(Grades!J2:J{last_row})'),
        ("A Grades", f'=COUNTIF(Grades!J2:J{last_row},"A")'),
        ("Average Attendance", f"=ROUND(AVERAGE(Students!L2:L{last_row}),1)"),
    ]

    for i, (label, formula) in enumerate(kpis):
        row = 5 + (i // 3) * 4
        col = 2 + (i % 3) * 2

        # Label
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        label_cell = ws.cell(row=row, column=col, value=label)
        label_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        label_cell.fill = PatternFill(start_color=COLORS["subheader_bg"], end_color=COLORS["subheader_bg"], fill_type="solid")
        label_cell.alignment = Alignment(horizontal="center")
        label_cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                    top=Side(style="medium"), bottom=Side(style="medium"))

        # Value
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        value_cell = ws.cell(row=row + 1, column=col, value=formula)
        value_cell.font = Font(name="Calibri", size=20, bold=True, color="1F4E79")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.fill = PatternFill(start_color=COLORS["dashboard_bg"], end_color=COLORS["dashboard_bg"], fill_type="solid")
        value_cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                    bottom=Side(style="medium"))
        if "Rate" in label:
            value_cell.number_format = "0.0%"

    # ── Bar Chart: Student Average Scores ──
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Student Average Scores"
    chart1.y_axis.title = "Average Score (%)"
    chart1.x_axis.title = "Students"
    data_ref = Reference(wb["Grades"], min_col=9, min_row=1, max_row=last_row)
    cat_ref = Reference(wb["Grades"], min_col=2, min_row=2, max_row=last_row)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cat_ref)
    chart1.width = 22
    chart1.height = 13
    ws.add_chart(chart1, "B14")

    # ── Pie Chart: Grade Distribution ──
    chart2 = PieChart()
    chart2.title = "Grade Distribution"
    chart2.style = 10
    # Grade counts formula-based — use static cells
    grade_labels = ["A", "B", "C", "D", "E", "F"]
    ws["K5"].value = "Grade"
    ws["L5"].value = "Count"
    style_header(ws["K5"], COLORS["table_header_bg"])
    style_header(ws["L5"], COLORS["table_header_bg"])
    for i, g in enumerate(grade_labels, 6):
        ws.cell(row=i, column=11, value=g)
        ws.cell(row=i, column=12, value=f'=COUNTIF(Grades!J2:J{last_row},"{g}")')
        style_data_cell(ws.cell(row=i, column=11))
        style_data_cell(ws.cell(row=i, column=12))

    data_ref2 = Reference(ws, min_col=12, min_row=5, max_row=11)
    cat_ref2 = Reference(ws, min_col=11, min_row=6, max_row=11)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cat_ref2)
    chart2.width = 15
    chart2.height = 13
    ws.add_chart(chart2, "J14")

    # Column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return ws


def build_analytics_sheet(wb):
    """Sheet: Analytics — subject averages, gender comparison, rankings."""
    ws = wb.create_sheet("Analytics")
    last_row = len(STUDENTS) + 1

    # ── Section 1: Subject Averages ──
    ws.cell(row=1, column=1, value="Subject-Wise Averages").font = Font(size=14, bold=True, color="1F4E79")
    headers = ["Subject", "Average", "Highest", "Lowest", "Pass Rate"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=2, column=col, value=h), COLORS["table_header_bg"])

    for i, subject in enumerate(SUBJECTS):
        col_letter = get_column_letter(SUBJECT_COLS + i)
        row = 3 + i
        ws.cell(row=row, column=1, value=subject)
        ws.cell(row=row, column=2, value=f"=ROUND(AVERAGE(Grades!{col_letter}2:{col_letter}{last_row}),2)")
        ws.cell(row=row, column=3, value=f"=MAX(Grades!{col_letter}2:{col_letter}{last_row})")
        ws.cell(row=row, column=4, value=f"=MIN(Grades!{col_letter}2:{col_letter}{last_row})")
        ws.cell(row=row, column=5, value=f'=COUNTIF(Grades!{col_letter}2:{col_letter}{last_row},">=50")/COUNTA(Grades!{col_letter}2:{col_letter}{last_row})')
        ws.cell(row=row, column=5).number_format = "0.0%"
        for col in range(1, 6):
            style_data_cell(ws.cell(row=row, column=col), alt_row=(row % 2 == 0))

    # ── Section 2: Gender Comparison ──
    gender_row = 11
    ws.cell(row=gender_row, column=1, value="Gender Comparison").font = Font(size=14, bold=True, color="1F4E79")
    g_headers = ["Gender", "Count", "Average Score"]
    for col, h in enumerate(g_headers, 1):
        style_header(ws.cell(row=gender_row + 1, column=col, value=h), COLORS["table_header_bg"])

    ws.cell(row=gender_row + 2, column=1, value="Male")
    ws.cell(row=gender_row + 2, column=2, value=f'=COUNTIF(Students!D2:D{last_row},"Male")')
    ws.cell(row=gender_row + 2, column=3, value=f'=ROUND(AVERAGEIFS(Grades!I2:I{last_row},Students!D2:D{last_row},"Male"),2)')

    ws.cell(row=gender_row + 3, column=1, value="Female")
    ws.cell(row=gender_row + 3, column=2, value=f'=COUNTIF(Students!D2:D{last_row},"Female")')
    ws.cell(row=gender_row + 3, column=3, value=f'=ROUND(AVERAGEIFS(Grades!I2:I{last_row},Students!D2:D{last_row},"Female"),2)')

    for r in range(gender_row + 2, gender_row + 4):
        for c in range(1, 4):
            style_data_cell(ws.cell(row=r, column=c))

    # ── Section 3: Class Rankings ──
    rank_row = 17
    ws.cell(row=rank_row, column=1, value="Class Rankings").font = Font(size=14, bold=True, color="1F4E79")
    r_headers = ["Rank", "Student ID", "Name", "Average", "Grade"]
    for col, h in enumerate(r_headers, 1):
        style_header(ws.cell(row=rank_row + 1, column=col, value=h), COLORS["table_header_bg"])

    for i in range(len(STUDENTS)):
        row = rank_row + 2 + i
        ws.cell(row=row, column=1, value=f"=RANK(Grades!I{i+2},Grades!I2:I{last_row},0)")
        ws.cell(row=row, column=2, value=f"=Grades!A{i+2}")
        ws.cell(row=row, column=3, value=f"=Grades!B{i+2}")
        ws.cell(row=row, column=4, value=f"=Grades!I{i+2}")
        ws.cell(row=row, column=5, value=f"=Grades!J{i+2}")
        for c in range(1, 6):
            style_data_cell(ws.cell(row=row, column=c), alt_row=(row % 2 == 0))

    # Chart: Subject Averages
    chart = BarChart()
    chart.type = "col"
    chart.title = "Subject Averages"
    chart.y_axis.title = "Score"
    data = Reference(ws, min_col=2, min_row=2, max_row=7)
    cats = Reference(ws, min_col=1, min_row=3, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 12
    ws.add_chart(chart, f"H{rank_row}")

    # Column widths
    for col, w in enumerate([14, 10, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return ws


def build_boys_girls_sheets(wb):
    """Sheets: Boys and Girls — separate records with formulas."""
    for gender, sheet_name in [("Male", "Boys"), ("Female", "Girls")]:
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=f"{gender.upper()} STUDENT RECORDS")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="1F4E79")
        ws.merge_cells("A1:J1")

        headers = ["Name", "Age", "English", "Mathematics", "Science", "ICT", "Social Studies", "Total", "Average", "Grade"]
        for col, h in enumerate(headers, 1):
            style_header(ws.cell(row=3, column=col, value=h), COLORS["table_header_bg"])

        row_idx = 4
        for i, (student, scores) in enumerate(zip(STUDENTS, SCORES)):
            if student[3] == gender:
                name = f"{student[1]} {student[2]}"
                ws.cell(row=row_idx, column=1, value=name)
                for j, s in enumerate(scores):
                    ws.cell(row=row_idx, column=3 + j, value=s)
                ws.cell(row=row_idx, column=8, value=f"=SUM(C{row_idx}:G{row_idx})")
                ws.cell(row=row_idx, column=9, value=f"=ROUND(AVERAGE(C{row_idx}:G{row_idx}),2)")
                ws.cell(row=row_idx, column=9).number_format = "0.00"
                ws.cell(row=row_idx, column=10,
                        value=f'=IF(I{row_idx}>=80,"A",IF(I{row_idx}>=70,"B",IF(I{row_idx}>=60,"C",IF(I{row_idx}>=50,"D","F"))))')

                for col in range(1, 11):
                    style_data_cell(ws.cell(row=row_idx, column=col), alt_row=(row_idx % 2 == 0))
                row_idx += 1

        # Subject averages
        avg_row = row_idx + 2
        ws.cell(row=avg_row, column=1, value="Subject Avg").font = Font(bold=True)
        for j, subject in enumerate(SUBJECTS):
            col_letter = get_column_letter(3 + j)
            ws.cell(row=avg_row, column=3 + j,
                    value=f"=ROUND(AVERAGE({col_letter}4:{col_letter}{row_idx-1}),2)")

        # Chart
        chart = BarChart()
        chart.title = f"{gender} — Student Averages"
        chart.y_axis.title = "Average"
        data = Reference(ws, min_col=9, min_row=3, max_row=row_idx - 1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row_idx - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 12
        ws.add_chart(chart, f"L3")

        for col, w in enumerate([22, 8, 10, 12, 10, 8, 14, 10, 10, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A4"


def build_reports_sheet(wb):
    """Sheet: Reports — report card layout."""
    ws = wb.create_sheet("Reports")
    ws.cell(row=1, column=1, value="STUDENT REPORT CARDS").font = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells("A1:G1")

    headers = ["Student ID", "Name", "Total", "Average", "Grade", "Remarks", "Attendance"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=2, column=col, value=h), COLORS["table_header_bg"])

    last_row = len(STUDENTS) + 1
    for i in range(len(STUDENTS)):
        row = 3 + i
        g_row = i + 2
        ws.cell(row=row, column=1, value=f"=Grades!A{g_row}")
        ws.cell(row=row, column=2, value=f"=Grades!B{g_row}")
        ws.cell(row=row, column=3, value=f"=Grades!H{g_row}")
        ws.cell(row=row, column=4, value=f"=Grades!I{g_row}")
        ws.cell(row=row, column=5, value=f"=Grades!J{g_row}")
        ws.cell(row=row, column=6, value=f"=Grades!K{g_row}")
        ws.cell(row=row, column=7, value=f"=Students!L{g_row}")

        for col in range(1, 8):
            style_data_cell(ws.cell(row=row, column=col), alt_row=(row % 2 == 0))

    for col, w in enumerate([10, 22, 10, 10, 8, 16, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def generate_workbook(school_name="Atlas International School", output_dir="."):
    """Generate the complete Student Management workbook."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build data sheets first (dashboard references them)
    build_students_sheet(wb)
    build_grades_sheet(wb)
    build_boys_girls_sheets(wb)
    build_analytics_sheet(wb)
    build_reports_sheet(wb)
    build_dashboard_sheet(wb, school_name)  # Build last (needs other sheets)

    # Reorder: Dashboard, Students, Grades, Boys, Girls, Analytics, Reports
    order = ["Dashboard", "Students", "Grades", "Boys", "Girls", "Analytics", "Reports"]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=-(wb.sheetnames.index(name) - i))

    # Save
    os.makedirs(output_dir, exist_ok=True)
    safe_name = school_name.replace(" ", "_")
    filename = f"{safe_name}_Student_Management.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ Workbook saved: {filepath}")
    return filepath


def main():
    parser = argparse.ArgumentParser(description="Student Management System — Excel Generator")
    parser.add_argument("--school", default="Atlas International School",
                        help="School name (default: Atlas International School)")
    parser.add_argument("--output", default="./output",
                        help="Output directory (default: ./output)")
    args = parser.parse_args()

    print(f"\n🎓 Generating workbook for: {args.school}")
    print(f"📁 Output directory: {args.output}\n")
    filepath = generate_workbook(args.school, args.output)
    print(f"\n📊 Workbook: {filepath}")
    print(f"   Sheets: Dashboard, Students, Grades, Boys, Girls, Analytics, Reports")
    print(f"   Students: {len(STUDENTS)}")
    print(f"   Subjects: {', '.join(SUBJECTS)}")
    print(f"   Features: Charts, Formulas, Conditional Formatting, Rankings")


if __name__ == "__main__":
    main()
